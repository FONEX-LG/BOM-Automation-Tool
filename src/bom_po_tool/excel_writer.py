from openpyxl import load_workbook
from models import PartLine


def generate_po_file(parts: list[PartLine], template_path: str, output_path: str):
    print(f"Generating PO: {output_path}...")

    try:
        wb = load_workbook(template_path)
        sheet = wb.worksheets[0]  # Always the first sheet
    except Exception as e:
        raise Exception(f"Could not open template file: {e}")

    header_row = None
    cols = {}

    print("Scanning for headers...")

    for r in range(1, 31):
        found_line_number = False

        # Look at every cell in this row
        for cell in sheet[r]:
            if not cell.value: continue

            # Convert to simple lowercase string
            txt = str(cell.value).lower()

            # We identify the row by finding "Line Number"
            if "line" in txt and "number" in txt and "item" not in txt:
                found_line_number = True
                header_row = r

        # Mapping
        if found_line_number:
            print(f"-> Found Header Row at {r}")
            for cell in sheet[r]:
                if not cell.value: continue
                txt = str(cell.value).lower()

                # Simple Keyword Mapping
                if "line" in txt and "number" in txt:
                    cols['line'] = cell.column
                elif "item" in txt and "number" in txt:
                    cols['item_num'] = cell.column
                elif "description" in txt:
                    cols['desc'] = cell.column
                elif "quantity" in txt:
                    cols['qty'] = cell.column
                elif "price" in txt:
                    cols['price'] = cell.column
                elif "supplier" in txt:
                    cols['supp_pn'] = cell.column
                elif "source" in txt:
                    cols['source'] = cell.column
                elif "category" in txt:
                    cols['cat'] = cell.column
                elif "comments" in txt:
                    cols['link'] = cell.column
                elif "need by" in txt:
                    cols['date'] = cell.column

            # Stop scanning rows once we found it
            break

    if not header_row:
        print("CRITICAL: 'Line Number' not found. Here is what I saw in the first 5 rows:")
        for r in range(1, 6):
            print(f"Row {r}: {[str(c.value) for c in sheet[r] if c.value]}")
        raise Exception("Could not find header row.")

    # Write Data
    current_row = header_row + 1
    line_count = 1

    for part in parts:
        if part.qty == 0: continue

        def write(key, val):
            if key in cols:
                sheet.cell(row=current_row, column=cols[key]).value = val

        write('line', line_count)
        write('qty', part.qty)
        write('price', part.unit_price)
        write('supp_pn', part.supplier_pn or part.mpn)

        # Description
        desc = part.description
        if part.manufacturer: desc += f" [{part.manufacturer}]"
        write('desc', desc)

        # Defaults
        write('source', "No Item")
        write('cat', "Lab Hardware under 1000$")
        write('link', part.link)

        # Empty Columns
        write('date', "")
        write('item_num', "")

        current_row += 1
        line_count += 1

    wb.save(output_path)
    print("Success.")